#!/usr/bin/python3
# encoding: utf-8 
# @Time    : 2018/4/14 0014 16:31
# @author  : zza
# @Email   : 740713651@qq.com
from openpyxl import load_workbook

###############读取
wb = load_workbook("./template.xlsx")

print(wb.sheetnames)  # [‘Sheet1‘, ‘Sheet2‘, ‘Sheet3‘]

sheet = wb.get_sheet_by_name("Sheet1")

c = sheet["C"]
print(c)
for i in range(len(c)):
    print(c[i].value)
c = sheet["4"]
print(c)
for i in range(len(c)):
    print(c[i].value)

print(sheet["C4"].value)  # c4     <-第C4格的值
print(sheet.max_row)  # 10     <-最大行数
print(sheet.max_column)  # 5     <-最大列数
for i in sheet["C"]:
    print(i.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值
###############写入

from openpyxl import Workbook

wb = Workbook()
sheet.title = "New Shit"
sheet['C3'] = 'Hello world!'
for i in range(10):
    sheet["A%d" % (i + 1)].value = i + 1

sheet["E1"].value = "=SUM(A:A)"

wb.save('保存一个新的excel.xlsx')
