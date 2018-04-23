#!/usr/bin/python3
# encoding: utf-8 
# @Time    : 2018/4/16 0016 10:58
# @author  : zza
# @Email   : 740713651@qq.com

import smtplib
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook


def load_email(email_list):
    wb = load_workbook(email_list)
    sheet = wb["Sheet1"]
    # 发送者的帐号密码
    admin = (sheet["A2"].value, sheet["B2"].value)
    #
    result = {}
    for i in range(4, sheet.max_row):
        name = sheet["A%d" % i].value
        email = sheet["B%d" % i].value
        if not name: continue
        if result.get(name):
            print("名字重复：", name)
            raise Exception()
        result.update({name: email})
    return admin, result


title = ""


def load_data(data_file):
    global title
    wb = load_workbook(data_file, data_only=True)
    sheet = wb["Sheet1"]
    result = {}
    for i in range(7, sheet.max_row):
        name = sheet["C" + str(i)].value
        datalist = []
        if not name: continue
        if result.get(name):
            print("名字重复：", name)
            raise Exception()
        for x in range(ord('a'), ord('z') + 1):
            data = str(sheet[chr(x) + str(i)].value)
            datalist.append(data if data else "")

        result.update({name: datalist})

    title = sheet["A2"].value
    return result


def get_msg(who, data):
    global title
    # 创建一个带附件的实例
    message = MIMEMultipart()
    message['From'] = Header("天河国云综合部", 'utf-8')
    message['To'] = Header(who, 'utf-8')
    subject = '天河国云个人工资条 - ' + title
    message['Subject'] = Header(subject, 'utf-8')
    message.attach(MIMEText(title, 'plain', 'utf-8'))

    msgAlternative = MIMEMultipart('alternative')
    message.attach(msgAlternative)
    msg = []
    msg.append("序号 : " + data[0])
    msg.append("部门 : " + data[1])
    msg.append("姓名 : " + data[2])
    msg.append("入职时间 : " + data[3])
    msg.append("职级 : " + data[4])

    msg.append("#" * 15 + "薪酬部分" + "#" * 15)
    msg.append("基本工资 : " + data[5])
    msg.append("保密工资 : " + data[6])
    msg.append("岗位工资 : " + data[7])
    msg.append("工龄工资 : " + data[8])
    msg.append("绩效工资 : " + data[9])
    msg.append("工资合计 : " + data[10])
    msg.append("补助 : " + data[11])
    msg.append("迟到早退 : " + data[12])
    msg.append("合计 : " + data[13])

    msg.append("#" * 15 + "应扣部分" + "#" * 15)
    msg.append("社保 : " + data[14])
    msg.append("公积金 : " + data[15])
    msg.append("小计 : " + data[16])
    msg.append("应发合计 : " + data[17])

    msg.append("#" * 15 + "其他扣款" + "#" * 15)
    msg.append("个税 : " + data[18])
    msg.append("年终奖个税 : " + data[19])
    msg.append("暂扣 : " + data[20])
    msg.append("合计 : " + data[21])

    msg.append("#" * 35)
    msg.append("补发部分 : " + data[22])
    msg.append("已发部分 : " + data[23])
    msg.append("实发工资 : " + data[24])
    msg.append("备注 : " + data[25])
    msg = "".join(["<p>" + i + "</p>\n" for i in msg])
    # 邮件正文内容
    msgAlternative.attach(MIMEText(msg, 'html', 'utf-8'))
    return message


def send_msg(admin, data_file, email_list):
    from_addr, password = admin
    receivers = email_list.values()  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
    # smtp_server = "smtp.qq.com"
    smtp_server = "smtp.ym.163.com"
    port = 25
    fales_list = []
    try:
        smtpObj = smtplib.SMTP(smtp_server, port)  # SMTP协议默认端口是25
        smtpObj.login(from_addr, password)
        for who in email_list.keys():
            try:
                if not who: continue
                message = get_msg(who, data_file[who])
                smtpObj.sendmail(from_addr, [email_list[who]], message.as_string())
                print(who + "\t\t\t" + "发送成功")
            except Exception as err:
                fales_list.append(who)
                print(who + "\t\t\t" + "发送失败")
                print(err)
    except smtplib.SMTPException:
        print("Error: 无法连接到邮箱服务器，请检查帐号密码\n",
              from_addr, password, smtp_server, port)
    return fales_list


def check(email_list, data):
    email_name = set(email_list.keys())
    data_name = set(data.keys())

    diff = data_name - email_name
    if len(diff) != 0:
        print("以下人员没有对应邮箱：\n", diff)

    diff = email_name - data_name
    if len(diff) != 0:
        print("以下人员没有工资项目：\n", diff)
    return diff


def main():
    try:
        email_file = "邮箱表.xlsx"
        # 获取email信息
        admin, email_list = load_email(email_file)
        print(email_file + "载入成功\n")

        # 获取工资信息
        data_file = "工资表.xlsx"
        data = load_data(data_file)
        print(data_file + "载入成功\n")
    except:
        print("邮箱表.xlsx 或 工资表.xlsx错误，请检查文件是否存在，格式是否正确")
        input("")
        return

    # 检查
    diff = check(email_list, data)

    input("按Enter键发送邮件")
    # 发送信息
    fales_list = send_msg(admin, data, email_list)
    print("\n邮件发送失败列表：")
    for i in fales_list:
        print(i)

    input("按enter结束")


if __name__ == '__main__':
    main()
